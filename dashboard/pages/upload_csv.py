import streamlit as st
import pandas as pd
import requests
import json
import io
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def header_name(ws, header_list, comment_list, width_list):
    for col_num, (headers, comments, widths) in enumerate(zip(header_list, comment_list, width_list), 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = headers
        cell.comment = Comment(comments, 'author', width=250, height=100)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = widths

def data_validation(ws, type, formula1, area):
    if type == 'list':
        dv = DataValidation(type=type, formula1=formula1, allow_blank=True)
        dv.promptTitle = 'List Selection'
        dv.prompt = 'Please select from the list'
        dv.errorTitle = 'Invalid Entry'
        dv.error = 'Your entry is not in the list'
    elif type == 'decimal':
        dv = DataValidation(type=type, operator='between', formula1=formula1, formula2='1', allow_blank=True)
        dv.promptTitle = 'Decimal Input'
        dv.prompt = 'Please input decimal number in range between 0 to 1'
        dv.errorTitle = 'Invalid Entry'
        dv.error = 'Your entry must be a decimal number between 0 and 1'
    elif type == 'whole':
        dv = DataValidation(type=type, operator='between', formula1=formula1, formula2='21', allow_blank=True)
        dv.promptTitle = 'Whole Number Input'
        dv.prompt = 'Please input whole number in range between 0 to 21'
        dv.errorTitle = 'Invalid Entry'
        dv.error = 'Your entry must be a whole number between 0 and 21'
    else:
        return
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(area)

def generate_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mass Input'
    header_list = ['Full Name', 'Gender', 'Enrolled University', 'Work Experience', 'Data Science Experience',
                   'Duration of Last New Job', 'Education Level', 'Major Discipline', 'City Development Index',
                   'Company Size', 'Company Type']
    comment_list = ['Enter the full name of the employee.',
                    'Select the gender of the employee.',
                    'Specify whether the employee is currently enrolled in a university.',
                    'Enter the total years of employee’s work experience. If more than 20 years, input ‘21’.',
                    'Specify if the employee has relevant experience in data science.',
                    'Enter the duration (in years) since the employee’s last new job.',
                    'Specify the highest education level achieved by the employee.',
                    'Specify the major discipline of the employee’s highest qualification.',
                    'Input the City Development Index for the location where the company is based.',
                    'Specify the size of the company based on the number of employees.',
                    'Enter the type of company.']
    width_list = [25] * len(header_list)
    header_name(ws, header_list, comment_list, width_list)
    data_validation(ws, "list", '"Male,Female,Other"', 'B2:B1001')
    data_validation(ws, "list", '"No Enroll,Part Time,Full Time"', 'C2:C1001')
    data_validation(ws, "whole", '0', 'D2:D1001')
    data_validation(ws, "list", '"Yes,No"', 'E2:E1001')
    data_validation(ws, "list", '"Never,1,2,3,4,More than 4"', 'F2:F1001')
    data_validation(ws, "list", '"Primary School,High School,Graduate,Masters,Phd"', 'G2:G1001')
    data_validation(ws, "list", '"STEM,Humanities,Business Degree,Arts,No Major,Other"', 'H2:H1001')
    data_validation(ws, "decimal", '0', 'I2:I1001')
    data_validation(ws, "list", '"Less than 10,10 to 49,50 to 99,100 to 499,500 to 999,1000 to 4999,5000 to 9999,More than 9999"', 'J2:J1001')
    data_validation(ws, "list", '"Pvt Ltd,Public Sector,Funded Startup,Early Startup,NGO,Other"', 'K2:K1001')
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def show(set_page):
    st.title("📤 Upload File XLSX untuk Prediksi Massal")
    template_file = generate_excel_template()
    st.download_button(
        label="📥 Download Template XLSX",
        data=template_file.getvalue(),
        file_name="Mass_Input_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    uploaded_file = st.file_uploader("Pilih file XLSX", type=["xlsx"])
    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        
        # Debugging: Tampilkan kolom yang ditemukan di dataframe
        # st.write("📄 **Kolom yang ditemukan dalam file Excel:**")
        # st.write(df.columns.tolist())  # Menampilkan daftar nama kolom
        
        # Normalisasi nama kolom untuk menghapus spasi ekstra
        df.columns = df.columns.str.strip()
        
        expected_columns = [
            "Full Name", "City Development Index", "Data Science Experience", "Enrolled University",
            "Education Level", "Work Experience", "Company Size", "Duration of Last New Job",
            "Gender", "Major Discipline", "Company Type"
        ]
        
          # Cek apakah semua kolom yang diharapkan ada dalam dataframe
        missing_columns = [col for col in expected_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Error: Kolom berikut tidak ditemukan dalam file Excel: {missing_columns}")
            return
        
        # Cek apakah ada data yang kosong di dalam dataframe
        if df.isnull().values.any():
            st.error("❌ Error: Terdapat data yang kosong dalam file Excel. Harap lengkapi semua kolom sebelum mengunggah.")
            return
        
        st.success("✅ Semua kolom sesuai, siap untuk diproses.")
        
        st.dataframe(df)  # Tampilkan dataframe yang di-load untuk validasi manual
        
        if st.button("🚀 Prediksi Semua Data"):
            url = "http://127.0.0.1:8000/predict"
            
            data = [
                {
                    "fullname": row["Full Name"],
                    "features": row[expected_columns[1:]].astype(str).tolist()  # Mengirim semua data mentah ke FastAPI
                }
                for _, row in df.iterrows()
            ]
            
            response = requests.post(url, json=data)
            
            if response.status_code == 200:
                predictions = response.json()["predictions"]
                df["Hasil Prediksi"] = ["💼 Mencari Pekerjaan Baru" if pred["prediction"] == 1 else "🏢 Tidak Mencari Pekerjaan Baru" for pred in predictions]
                st.success("✅ Prediksi Berhasil!")
                st.dataframe(df)
            else:
                st.error(f"Error: {response.text}")
    if st.button("⬅️ Kembali ke Dashboard"):
        set_page("dashboard")
