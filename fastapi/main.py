from fastapi import FastAPI, HTTPException, Depends
from fastapi.responses import StreamingResponse, HTMLResponse
from pyngrok import ngrok
import uvicorn
from pydantic import BaseModel, Field
from enum import Enum
from typing import List, Dict, Any
import pandas as pd
import joblib
from sklearn.preprocessing import OrdinalEncoder, MinMaxScaler
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import Workbook, utils
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment
from io import BytesIO
from langchain_ollama import OllamaLLM
from langchain_experimental.agents import create_pandas_dataframe_agent
from dotenv import load_dotenv
import os
from datetime import datetime
from functools import lru_cache

# Load environment variables
load_dotenv()
ngrok_auth_token = os.getenv('NGROK_AUTH_TOKEN')
if not ngrok_auth_token:
  raise Exception("NGROK_AUTH_TOKEN not found in environment variables")

  
# Expose the FastAPI app with ngrok
ngrok.set_auth_token(ngrok_auth_token)

# Create FastAPI app
app = FastAPI(
    title="Employee Prediction API",
    description="API for predicting employee job change after course completion",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Load pickle
try:
    ordinalencoder = joblib.load(os.path.join('pickle', 'ordinalencoder.pkl'))
    minmaxscaler = joblib.load(os.path.join('pickle', 'minmaxscaler.pkl'))
    model = joblib.load(os.path.join('pickle', 'lclgbm.pkl'))
except Exception as e:
  raise Exception("Error loading pickle")

# Class
## Class: categorical columns
class gender_cat(str, Enum):
  male = 'Male'
  female = 'Female'
  other = 'Other'

class enrolled_university_cat(str, Enum):
  no_enroll = 'No Enroll'
  part_time = 'Part Time'
  full_time = 'Full Time'

class education_level_cat(str, Enum):
  primary_school = 'Primary School'
  high_school = 'High School'
  graduate = 'Graduate'
  masters = 'Masters'
  phd = 'Phd'

class major_discipline_cat(str, Enum):
  stem = 'STEM'
  humanities = 'Humanities'
  business_degree = 'Business Degree'
  arts = 'Arts'
  no_major = 'No Major'
  other = 'Other'

class experience_cat(str, Enum):
  zero_exp = '<1'
  one_exp = '1'
  two_exp = '2'
  three_exp = '3'
  four_exp = '4'
  five_exp = '5'
  six_exp = '6'
  seven_exp = '7'
  eight_exp = '8'
  nine_exp = '9'
  ten_exp = '10'
  eleven_exp = '11'
  twelve_exp = '12'
  thirteen_exp = '13'
  fourteen_exp = '14'
  fifteen_exp = '15'
  sixteen_exp = '16'
  seventeen_exp = '17'
  eighteen_exp = '18'
  nineteen_exp = '19'
  twenty_exp = '20'
  more_exp = '>20'

class company_size_cat(str, Enum):
  less_10 = '<10'
  around_10_49 = '10-49'
  around_50_99 = '50-99'
  around_100_499 = '100-500'
  around_500_999 = '500-999'
  around_1000_4999 = '1000-4999'
  around_5000_9999 = '5000-9999'
  more_10000 = '10000+'

class company_type_cat(str, Enum):
  pvt_ltd = 'Pvt Ltd'
  public_sector = 'Public Sector'
  funded_startup = 'Funded Startup'
  early_startup = 'Early Startup'
  ngo = 'NGO'
  other = 'Other'

class last_new_job_cat(str, Enum):
  never = 'never'
  one = '1'
  two = '2'
  three = '3'
  four = '4'
  more_four = '>4'

## Class: Combine into one input
class EmployeeData(BaseModel):
  full_name: str = Field(..., max_length = 200)
  city_development_index: float = Field(..., ge = 0, le = 1)
  gender: gender_cat
  relevant_experience: bool = Field(...)
  enrolled_university: enrolled_university_cat
  education_level: education_level_cat
  major_discipline: major_discipline_cat
  experience: experience_cat
  company_size: company_size_cat
  company_type: company_type_cat
  last_new_job: last_new_job_cat

  class Config:
    json_schema_extra = {
      "example": {
        "full_name": "John Doe",
        "city_development_index": 0.5,
        "gender": "Male",
        "relevant_experience": True,
        "enrolled_university": "No Enroll",
        "education_level": "Graduate",
        "major_discipline": "STEM",
        "experience": "1",
        "company_size": "10-49",
        "company_type": "Public Sector",
        "last_new_job": "1"
      }
    }

## Class: Mass input
class MassInputData(BaseModel):
  employees: List[EmployeeData] = Field(..., max_items = 50)

## Class: Preprocessed Data
class PreprocessedData(BaseModel):
  original_data: List[Dict[str,Any]]
  preprocessed_features: List[Dict[str,float]]
  features_columns: List[str]

## Class: Input to LLM AI
class AIRequest(BaseModel):
  question: str
  df_dict: dict

## Class: Success Response from LLM AI
class SuccesResponse(BaseModel):
  status: str = "success"
  message: str
  by: str

## Class: Error Response from LLM AI
class ErrorResponse(BaseModel):
  status: str = "error"
  message: str

# Map and Column
## gender
gender_map = {
    "Female": [1, 0],
    "Male": [0, 1],
    "Other": [0, 0]
    }
gender_columns = [
    "gender_Female",
    "gender_Male"
    ]
## major_discipline
major_discipline_map = {
    "Arts": [1, 0, 0, 0, 0],
    "Business Degree": [0, 1, 0, 0, 0],
    "Humanities": [0, 0, 1, 0, 0],
    "No Major": [0, 0, 0, 1, 0],
    "STEM": [0, 0, 0, 0, 1],
    "Other": [0, 0, 0, 0, 0]
  }
major_discipline_columns = [
    "major_discipline_Arts",
    "major_discipline_Business Degree",
    "major_discipline_Humanities",
    "major_discipline_No Major",
    "major_discipline_STEM"
  ]
## company_type
company_type_map = {
    "Early Startup": [1, 0, 0, 0, 0],
    "Funded Startup": [0, 1, 0, 0, 0],
    "NGO": [0, 0, 1, 0, 0],
    "Public Sector": [0, 0, 0, 1, 0],
    "Pvt Ltd": [0, 0, 0, 0, 1],
    "Other": [0, 0, 0, 0, 0]
  }
company_type_columns = [
    "company_type_Early Startup",
    "company_type_Funded Startup",
    "company_type_NGO",
    "company_type_Public Sector",
    "company_type_Pvt Ltd"
  ]


# Func: Create Excel Template
## Sub-Func: Create Header
def header_name(ws,header_list,comment_list,width_list):
  for col_num, (headers,comments,widths) in enumerate(zip(header_list,comment_list,width_list),1):
    # initiate cell
    cell = ws.cell(row=1,column=col_num)
    # fill headers with text
    cell.value = headers
    # fill headers with comment
    cell.comment = Comment(comments, 'author', width=250, height=100)
    # style headers
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')
    # resize headers
    column_letter = utils.get_column_letter(col_num)
    ws.column_dimensions[column_letter].width = widths

## Sub-Func: Create Data Validation
def data_validation(ws,type,formula1,area):
  if type == 'list':
    dv = DataValidation(type=type, formula1=formula1)
    # pop up for instruction
    dv.promptTitle = 'List Selection'
    dv.prompt = 'Please select from the list'
    # if input not in data validation criteria
    dv.errorTitle = 'Invalid Entry'
    dv.error ='Your entry is not in the list'

  elif type == 'decimal':
    dv = DataValidation(type=type, operator='between',formula1=formula1, formula2=1)
    # pop up for instruction
    dv.promptTitle = 'Decimal Input'
    dv.prompt = 'Please input decimal number in range between 0 to 1'
    # if input not in data validation criteria
    dv.errorTitle = 'Invalid Entry'
    dv.error ='Your entry must decimal number in range between 0 to 1'

  elif type == 'whole':
    dv = DataValidation(type=type, operator='between',formula1=formula1, formula2=21)
    # pop up for instruction
    dv.promptTitle = 'Whole Number Input'
    dv.prompt = 'Please input whole number in range between 0 to 21'
    # if input not in data validation criteria
    dv.errorTitle = 'Invalid Entry'
    dv.error ='Your entry must be whole number in range between 0 to 21'
  
  elif type == 'textLength':
    dv = DataValidation(type=type, operator='lessThan',formula1=formula1)
    # pop up for instruction
    dv.promptTitle = 'Text Input'
    dv.prompt = 'Please input text with less than 200 characters'
    # if input not in data validation criteria
    dv.errorTitle = 'Invalid Entry'
    dv.error ='Your entry must be text with less than 200 characters'

  else:
    print('Invalid type')

  # show error and input message
  dv.showErrorMessage = True
  dv.showInputMessage = True
  # add data validation
  ws.add_data_validation(dv)
  dv.add(area)

## Main-Func: Excel
def generate_excel_template():
  # initiate excel
  wb = Workbook()
  ws = wb.active
  ws.title = 'Mass Input'
  # initiate header, comment, and width list
  header_list = ['Full Name','Gender','Enrolled University','Work Experience','Data Science Experience','Duration of Last New Job','Education Level','Major Discipline',
                'City Development Index','Company Size','Company Type']
  comment_list = [
      'Enter the full name of the employee.',
      'Select the gender of the employee.',
      'Specify whether the employee is curently enroll in a university.',
      'Enter the total years of employee’s work experience. If more than 20 years, input ’21’.',
      'Specify if employee have relevant experience in data science.',
      'Enter the duration (in years) since the employee’s last new job.',
      'Specify the highest education level achieved by the employee.',
      'Specify the major discipline of the employee’s highest qualification. If the employee’s highest education level is High School or lower, input ’No Major’.',
      'Input the City Development Index (a metric representing the level of urban development) for the location where the company is based.',
      'Specify the size of the company based on the number of employees.',
      'Enter the type of company.'
  ]
  width_list = [25] * 11
  # initiate headers
  header_name(ws,header_list,comment_list,width_list)
  # add data validation to each columns
  # a. fullname
  data_validation(ws,"textLength",200,'A2:A51')
  # b. gender
  data_validation(ws,"list",'"Male,Female,Other"','B2:B51')
  # c. enrolled_university
  data_validation(ws,"list",'"No Enroll,Part Time,Full Time"','C2:C51')
  # d. experience
  data_validation(ws,"whole",0,'D2:D51')
  # e. relevant_experience
  data_validation(ws,"list",'"Yes,No"','E2:E51')
  # f. last_new_job
  data_validation(ws,"list",'"Never,1,2,3,4,More than 4"','F2:F51')
  # g. education_level
  data_validation(ws,"list",'"Primary School,High School,Graduate,Masters,Phd"','G2:G51')
  # h. major_discipline
  data_validation(ws,"list",'"STEM,Humanities,Business Degree,Arts,No Major,Other"','H2:H51')
  # i. city_development_index
  data_validation(ws,"decimal",0,'I2:I51')
  # j. company_size
  data_validation(ws,"list",'"Less than 10,10 to 49,50 to 99,100 to 499,500 to 999,1000 to 4999,5000 to 9999,More than 9999"','J2:J51')
  # k. company_type
  data_validation(ws,"list",'"Pvt Ltd,Public Sector,Funded Startup,Early Startup,NGO,Other"','K2:K51')
  # Save the workbook to an in-memory file
  file_stream = BytesIO()
  wb.save(file_stream)
  file_stream.seek(0)
  return file_stream

# Func: LLM AI
## Sub-Func: Crate prompt prefix
@lru_cache(maxsize=1)
def get_prefix():
  return """
  You are an AI Assistant for Ascencio, a data science agency.
  Ascencio, a leading Data Science agency, offers training courses to companies to enhance their employees' skills.
  Companies want to predict which employees are unlikely to seek a job change after completing the course.
  By focusing on employees who are committed to staying, Ascencio helps companies optimize their training investments.
  To achieve this, Ascencio builds machine learning models based on LGBMClassifier to predict if employees will be looking for a job change or not after a data science course.
  This machine learning models will result in dataframe with following columns and description:
  - Full Name = Employee's full name
  - Gender = Employee's gender
  - Enrolled University = Specify whether the employee is curently enroll in a university
  - Work Experience = Employee's work experience in years
  - Data Science Experience = Specify if employee have relevant experience in data science
  - Duration of Last New Job = Duration since the employee's last new job in years
  - Education Level = Specify the highest education level achieved by the employee
  - Major Discipline = Employee's major field of study
  - City Development Index = City Development Index (a metric representing the level of urban development) for the location where the company is based
  - Company Size = Specify the size of the company based on the number of employees
  - Company Type = Type of company
  - Probability of Leaving = Probability of employee leaving the company after data science course
  - Prediction = Prediction of employee leaving the company after data science course

  If you need to choose employee for participate data science, you should prioritize employe with lowest probability of leaving 

  Column name always have space between words, example Full Name is name column and not FullName

  In following format below,
  When you input in 'Action Input:', never add backticks "`" around the action input and don't make new dataframe using pd.DataFrame

  When you give 'Final Answer:', never give suggestion about python and about code in 'Action Input:' and only give data reasoning analysis and give next step reccomendation for company
  """

## Main-Func: LLM AI Agent
def create_agent(df, model_name="qwen2.5", temp=0, max_execution_time=60):
  llm = OllamaLLM(model=model_name, temperature=temp)
  agent = create_pandas_dataframe_agent(
    llm,
    df,
    prefix=get_prefix(),
    number_of_head_rows=51,
    verbose=True,
    allow_dangerous_code=True,
    max_execution_time=max_execution_time
  )
  return agent

@app.get("/")
async def read_root():
    """Check if API is running and pickle files are loaded"""
    return {
      "status": "API running",
      "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
      "pickle_files": all([ordinalencoder, minmaxscaler, model])
      }

@app.get("/create_excel_template")
async def create_excel_template():
    """Generate and return an Excel template for mass input"""
    try:
        excel_file = generate_excel_template()
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=mass_input_template.xlsx"}
        )
    except Exception as e:
      raise HTTPException(
        status_code=500, 
        detail="Failed to generate Excel template"
      )

@app.post("/preprocess", response_model=PreprocessedData)
async def preprocess_data(data: MassInputData):
  """Preprocess employee(s) data for prediction (encoding and scaling)"""
  try:
    # Convert to dataframe
    df = pd.DataFrame([item.model_dump() for item in data.employees])
    # Store original data
    original_data = df.to_dict('records')
    # Divide columns
    oe_columns = ['relevant_experience', 'enrolled_university', 'education_level', 'experience', 'company_size', 'last_new_job']
    ohe_columns = ['gender','major_discipline','company_type']
    numerical_columns = ['city_development_index']
    # Ordinal Encoder oe_columns
    df[oe_columns] = ordinalencoder.transform(df[oe_columns])
    # One Hot Encoding ohe_columns
    gender_values = df['gender'].map(gender_map).tolist()
    major_discipline_values = df['major_discipline'].map(major_discipline_map).tolist()
    company_type_values = df['company_type'].map(company_type_map).tolist()
    gender_df = pd.DataFrame(gender_values, columns=gender_columns)
    major_discipline_df = pd.DataFrame(major_discipline_values, columns=major_discipline_columns)
    company_type_df = pd.DataFrame(company_type_values, columns=company_type_columns)
    df = pd.concat([df, gender_df, major_discipline_df, company_type_df], axis=1)
    # Drop unwanted columns
    df = df.drop(columns=ohe_columns)
    df = df.drop(columns='full_name')
    # Min Max Scaler
    df_values = minmaxscaler.transform(df)
    df = pd.DataFrame(df_values, columns=df.columns)
    # Store preprocessed features
    preprocessed_features = df.to_dict('records')
    # Store features columns
    features_columns = df.columns.tolist()
    return PreprocessedData(
        original_data=original_data,
        preprocessed_features=preprocessed_features,
        features_columns=features_columns
        )
  except Exception as e:
    raise HTTPException(
        status_code=500,
        detail=f"Error in preprocessing: {str(e)}"
    )

@app.post("/predict")
async def predict_data(data: PreprocessedData):
  """Make predictions based on preprocessed data"""
  try:
    # Convert PreprocessedData to dataframe
    df_final = pd.DataFrame(
              data.preprocessed_features,
              columns=data.features_columns
          )
    # Make predictions using model
    predictions = model.predict(df_final)
    probabilities = model.predict_proba(df_final)[:, 1]
    # Combine predictions with original data
    results = []
    for orig, pred, prob in zip(data.original_data, predictions, probabilities):
      results.append(
          {
              "original_data": orig,
              "prediction": pred,
              "probability": prob
          }
      )
    return {
        'status': 'success',
        'results': results
    }
  except Exception as e:
    raise HTTPException(
        status_code=500,
        detail=f"Error in prediction: {str(e)}"
    )
  
@app.post("/ai_ask", response_model=SuccesResponse, responses={500: {"model": ErrorResponse}})
async def ai_ask(request: AIRequest):
  try:
    # rearrange dataframe
    df = pd.DataFrame.from_dict(request.df_dict)

    # first try with qwen2.5
    agent = create_agent(df)
    result = agent.invoke(request.question)

    if result['output'] == 'Agent stopped due to iteration limit or time limit.':
      # second try with llama3.1
      agent = create_agent(df, model_name="llama3.1")
      result = agent.invoke(request.question)

      if result['output'] == 'Agent stopped due to iteration limit or time limit.':
        return ErrorResponse(
          message = "Analysis timed out. Please simplify your question or try again later."
        )
    
      return SuccesResponse(
        message=result['output'],
        by='Generated by Llama3.1'
      )
    
    return SuccesResponse(
      message=result['output'],
      by='Generated by Qwen2.5'
    )
  
  except Exception as e:
    raise HTTPException(
      status_code=500,
      detail=f"Error in LLM: {str(e)}"
    ) 
    
if __name__ == "__main__":
    # Connect to ngrok
    url = ngrok.connect(8000)
    print(f"Public URL: {url}")

    # Run the app with Uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)