from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
import joblib
import numpy as np
import logging
from typing import List
from fastapi import FastAPI, Response
import io
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Inisialisasi FastAPI
app = FastAPI()

# Load model XGBClassifier
model = joblib.load('app/model/model.pkl')

# Logging untuk debugging
logging.basicConfig(level=logging.INFO)

# Mapping kategori ke angka
enrolled_university_map = {"No Enroll": 0, "Part Time": 1, "Full Time": 2}
relevant_experience_map = {"No": 0, "Yes": 1}
education_level_map = {"Primary School": 0, "High School": 1, "Graduate": 2, "Masters": 3, "Phd": 4}
company_size_map = {
    "Less than 10": 0, "10 to 49": 1, "50 to 99": 2, "100 to 499": 3,
    "500 to 999": 4, "1000 to 4999": 5, "5000 to 9999": 6, "More than 9999": 7
}

gender_map = {"Male": [0, 1], "Female": [1, 0], "Other": [0, 0]}
major_discipline_map = {
    "Arts": [1, 0, 0, 0, 0],
    "Business Degree": [0, 1, 0, 0, 0],
    "Humanities": [0, 0, 1, 0, 0],
    "No Major": [0, 0, 0, 1, 0],
    "STEM": [0, 0, 0, 0, 1]
}
company_type_map = {
    "Early Startup": [1, 0, 0, 0, 0],
    "Funded Startup": [0, 1, 0, 0, 0],
    "NGO": [0, 0, 1, 0, 0],
    "Public Sector": [0, 0, 0, 1, 0],
    "Pvt Ltd": [0, 0, 0, 0, 1]
}

# Skema data untuk satu input
class InputData(BaseModel):
    fullname: str
    features: List[str]  # Semua fitur dalam bentuk string

# Endpoint untuk cek status API
@app.get("/")
def read_root():
    return {"message": "API LGBMClassifier berjalan dengan baik!"}

# Fungsi untuk encoding data kategori menjadi numerik
def encode_features(raw_features: List[str]) -> List[float]:
    try:
        logging.info(f"Raw features received: {raw_features}")

        # Konversi fitur numerik
        city_development_index = float(raw_features[0])
        relevant_experience = relevant_experience_map.get(raw_features[1], 0)
        enrolled_university = enrolled_university_map.get(raw_features[2], 0)
        education_level = education_level_map.get(raw_features[3], 0)
        experience = int(raw_features[4])
        company_size = company_size_map.get(raw_features[5], 0)
        last_new_job = int(raw_features[6])

        # Encoding gender
        gender_value = raw_features[7]
        gender_encoded = gender_map.get(gender_value, [0, 0])

        # Encoding major_discipline
        major_discipline_value = raw_features[8]
        major_discipline_encoded = major_discipline_map.get(major_discipline_value, [0, 0, 0, 0, 0])

        # Encoding company_type
        company_type_value = raw_features[9]
        company_type_encoded = company_type_map.get(company_type_value, [0, 0, 0, 0, 0])

        # Gabungkan semua fitur menjadi satu list
        encoded_features = [
            city_development_index, relevant_experience, enrolled_university, education_level,
            experience, company_size, last_new_job
        ] + gender_encoded + major_discipline_encoded + company_type_encoded

        logging.info(f"Encoded features: {encoded_features}")
        return encoded_features
    except Exception as e:
        raise ValueError(f"Error encoding features: {e}")

# Endpoint untuk prediksi banyak data
@app.post("/predict")
def predict(data: List[InputData]):
    try:
        logging.info(f"Request data: {data}")

        # Encoding semua data input menjadi format numerik
        features = np.array([encode_features(item.features) for item in data])

        # Validasi jumlah fitur
        expected_feature_count = 19
        if features.shape[1] != expected_feature_count:
            raise ValueError(f"Expected {expected_feature_count} features, got {features.shape[1]}")

        # Prediksi untuk semua data
        predictions = model.predict(features)

        return {
            "predictions": [
                {"fullname": item.fullname, "prediction": int(pred)}
                for item, pred in zip(data, predictions)
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error saat memproses data: {e}")




def header_name(ws, header_list, comment_list, width_list):
    for col_num, (headers, comments, widths) in enumerate(zip(header_list, comment_list, width_list), 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = headers
        cell.comment = Comment(comments, 'author', width=250, height=100)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = widths

def data_validation(ws, type, formula1, area):
    if type == 'list':
        dv = DataValidation(type=type, formula1=formula1, allow_blank=True)
        dv.promptTitle = 'List Selection'
        dv.prompt = 'Please select from the list'
        dv.errorTitle = 'Invalid Entry'
        dv.error = 'Your entry is not in the list'
    elif type == 'decimal':
        dv = DataValidation(type=type, operator='between', formula1=formula1, formula2='1', allow_blank=True)
        dv.promptTitle = 'Decimal Input'
        dv.prompt = 'Please input decimal number in range between 0 to 1'
        dv.errorTitle = 'Invalid Entry'
        dv.error = 'Your entry must be a decimal number between 0 and 1'
    elif type == 'whole':
        dv = DataValidation(type=type, operator='between', formula1=formula1, formula2='21', allow_blank=True)
        dv.promptTitle = 'Whole Number Input'
        dv.prompt = 'Please input whole number in range between 0 to 21'
        dv.errorTitle = 'Invalid Entry'
        dv.error = 'Your entry must be a whole number between 0 and 21'
    else:
        return
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(area)

    
def generate_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mass Input'

    header_list = ['Full Name', 'Gender', 'Enrolled University', 'Work Experience', 'Data Science Experience',
                   'Duration of Last New Job', 'Education Level', 'Major Discipline', 'City Development Index',
                   'Company Size', 'Company Type']
    comment_list = ['Enter the full name of the employee.',
                    'Select the gender of the employee.',
                    'Specify whether the employee is currently enrolled in a university.',
                    'Enter the total years of employee’s work experience. If more than 20 years, input ‘21’.',
                    'Specify if the employee has relevant experience in data science.',
                    'Enter the duration (in years) since the employee’s last new job.',
                    'Specify the highest education level achieved by the employee.',
                    'Specify the major discipline of the employee’s highest qualification.',
                    'Input the City Development Index for the location where the company is based.',
                    'Specify the size of the company based on the number of employees.',
                    'Enter the type of company.']
    width_list = [25] * len(header_list)

    for col_num, (headers, comments, widths) in enumerate(zip(header_list, comment_list, width_list), 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = headers
        cell.comment = Comment(comments, 'author', width=250, height=100)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = widths

    data_validation(ws, "list", '"Male,Female,Other"', 'B2:B1001')
    data_validation(ws, "list", '"No Enroll,Part Time,Full Time"', 'C2:C1001')
    data_validation(ws, "whole", '0', 'D2:D1001')
    data_validation(ws, "list", '"Yes,No"', 'E2:E1001')
    data_validation(ws, "list", '"Never,1,2,3,4,More than 4"', 'F2:F1001')
    data_validation(ws, "list", '"Primary School,High School,Graduate,Masters,Phd"', 'G2:G1001')
    data_validation(ws, "list", '"STEM,Humanities,Business Degree,Arts,No Major,Other"', 'H2:H1001')
    data_validation(ws, "decimal", '0', 'I2:I1001')
    data_validation(ws, "list", '"Less than 10,10 to 49,50 to 99,100 to 499,500 to 999,1000 to 4999,5000 to 9999,More than 9999"', 'J2:J1001')
    data_validation(ws, "list", '"Pvt Ltd,Public Sector,Funded Startup,Early Startup,NGO,Other"', 'K2:K1001')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.get("/download-template")
def download_template():
    """Endpoint untuk mengunduh template Excel."""
    file = generate_excel_template()
    headers = {
        "Content-Disposition": "attachment; filename=Mass_Input_Template.xlsx"
    }
    return Response(file.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
